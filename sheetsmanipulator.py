from openpyxl import load_workbook



wb=load_workbook(filename='modelo.xlsx')
ws=wb.active

nome='Eduardo'
ws['B1']=nome

wb.save("teste.xlsx")